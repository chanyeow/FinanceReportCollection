import openpyxl
import math
import os
import time
import requests
import xlrd
import platform
import os.path
import sys

from pathlib import Path

#ERROR_DEFINE
ERROR_STOCKLIST_NOTFOUND = 1
ERROR_STOCKLIST_SHEETNOTFOUND = 2
ERROR_WORKBOOK_SHEETNOTFOUND = 3

RESPONSE_TIMEOUT = 10
YEAR_LIST = [2017, 2018, 2019, 2020, 2021]


URL= 'http://90.push2his.eastmoney.com/api/qt/stock/kline/get'
HEADER = {
    #'User-Agent': 'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.36',
    #'X-Requested-With': 'XMLHttpRequest'
    'User-Agent: Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/70.0.3538.25 Safari/537.36 Core/1.70.3878.400 QQBrowser/10.8.4518.400',
    'X-Requested-With: XMLHttpRequest'
}

QUERY = {
    #'cb' : 'jQuery112402816109881259974_1638893884237',
    'secid' : '0.000001',
    'ut' : 'fa5fd1943c7b386f172d6893dbfba10b',
    'fields1' : 'f1,f2,f3,f4,f5,f6',
    'fields2' : 'f51,f52,f53,f54,f55,f56,f57,f58,f59,f60,f61',
    'klt' : 103,
    'fqt' : 0,
    'beg' : '20170101',
    'end' : '20500101',
    'smplmt' : 460,
    'lmt' : 1000000,
    '_' : '1638893884280',
}

class Stock:
#"2017-01-26,9.11,9.33,9.34,9.07,7629259,7001209275.00,2.97,2.53,0.23,4.51",
#时间，开盘价，收盘价，最高价，最低价
    def __init__(self, code, name, data):
        self.code = code
        self.name = name
        self.data = data
        self.HigherPrice = {}
        self.LowerPrice = {}

    def get_base_info(self):
        return self.code, self.name

    def get_Higher_Lower_Price(self, year):
        if self.HigherPrice and self.LowerPrice:
            if self.HigherPrice.__contains__(year) and self.LowerPrice.__contains__(year):
                return self.HigherPrice.get(year, 0), self.LowerPrice.get(year, 0)
        
        iLastHigherPrice = 0
        iLastLowerPrice = 0
        for each in self.data:
            sYear = each[:4]
            if (year == int(sYear)):
                strlist = each.split(',')
                iCurHigherPrice = float(strlist[3])
                if iCurHigherPrice > iLastHigherPrice:
                    iLastHigherPrice = iCurHigherPrice
                iCurLowerPrice = float(strlist[4])
                if iCurLowerPrice > iLastLowerPrice:
                    iLastLowerPrice = iCurLowerPrice

        self.HigherPrice[year] = iLastHigherPrice
        self.LowerPrice[year] = iLastLowerPrice
        return iLastHigherPrice, iLastLowerPrice

def get_code_secid(code):
    if code != '':
        if code[0] == '6':
            marketid = 1
        else:
            marketid = 0
        return '%d.%s'%(marketid, code)
    else:
        return ''

FILE_NAME_STOCK_LIST = 'stocklist.xlsx'
def get_all_stock_code():
    mData = {}
    #load file
    file = Path(FILE_NAME_STOCK_LIST)
    if file.is_file():
        stockList = openpyxl.load_workbook(FILE_NAME_STOCK_LIST)  
    else:
        return mData

    #get sheet name
    sheets = stockList.get_sheet_names()
    stockListSheet = stockList.get_sheet_by_name(sheets[0])
    #最大行数
    max_row = stockListSheet.max_row 
    #最大列数
    max_column = stockListSheet.max_column 

    for m in range(1, max_row + 1):
        i ='%s%d'%('a', m)
        key = stockListSheet[i].value
        i ='%s%d'%('b', m)
        value = stockListSheet[i].value
        mData[key] = value
    return mData


PROXIES = {
    # "http": "http://101.132.189.87:9090", 
    "http": "http://101.133.239.96:8080", 
}

def spidering(secid):
    QUERY['secid'] = secid
    try:
        # r = requests.post(URL, QUERY, HEADER, proxies=PROXIES, timeout=RESPONSE_TIMEOUT)
        r = requests.post(URL, QUERY, HEADER, timeout=RESPONSE_TIMEOUT)
    except Exception as e:
        print(e)
    if r.status_code == requests.codes.ok and r.text != '':
        data = r.json()
    else:
        data = {}

    try:
        r.close()
    except Exception as e:
        print(e)
    return data

def spideringAll(mData):
    mAllStockData = {}
    iTotalCount = len(mData)
    iIndex = 0
    for code, name in mData.items():
        iIndex += 1
        secid = get_code_secid(code)
        allData = spidering(secid)
        kData = allData['data']
        objStock = Stock(code, name, kData['klines'])
        for year in YEAR_LIST:
            objStock.get_Higher_Lower_Price(year)

        mAllStockData[code] = objStock
        print("正在爬取 %s - %s 的数据(%d/%d)"%(code, name, iIndex, iTotalCount))
        time.sleep(0.01)

    return mAllStockData
    
FILE_NAME_OUTPUT = '股票.xlsx'
TITLE_LIST = ['代码', '名称', 
    '17年最高价', '17年最低价',
    '18年最高价', '18年最低价',
    '19年最高价', '19年最低价',
    '20年最高价', '20年最低价',    
    '21年最高价', '21年最低价',

    '17与18幅度', '17与19幅度', '17与20幅度', '17与21幅度',
    '18与19幅度', '18与20幅度', '18与21幅度',
    '19与20幅度', '19与21幅度',
    '20与21幅度'
]

def main():
    mAllStockData = spideringAll(get_all_stock_code())

    print('数据爬取完毕，开始处理数据...')
    file = Path(FILE_NAME_OUTPUT)
    if not file.is_file():
        wb = openpyxl.Workbook()
        wb.save(FILE_NAME_OUTPUT)
    wb = openpyxl.load_workbook(FILE_NAME_OUTPUT)
    allSheest = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(allSheest[0])
    ws.append(TITLE_LIST)
    
    for code, objStock in mAllStockData.items():
        stockData = []

        #base info
        code, name = objStock.get_base_info()
        stockData.append(code)
        stockData.append(name)
        
        #price
        for year in YEAR_LIST:
            iHigherPrice, iLowerPrice = objStock.get_Higher_Lower_Price(year)
            stockData.append(iHigherPrice)
            stockData.append(iLowerPrice)

        #advanced
        for year in YEAR_LIST:
            iCurYearHigherPrice, iCurYearLowerPrice = objStock.get_Higher_Lower_Price(year)
            for y in YEAR_LIST:
                if (year >= y):
                    continue
                iHigherPrice, iLowerPrice = objStock.get_Higher_Lower_Price(y)

                iPrice = iHigherPrice - iCurYearLowerPrice
                if iCurYearLowerPrice > 0:
                    stockData.append(iPrice * 100 / iCurYearLowerPrice)
                else:
                    stockData.append(0)

        ws.append(stockData)

    wb.save(FILE_NAME_OUTPUT)
    wb.close
    print('数据全部处理完毕！')

if __name__ == '__main__':
    main()