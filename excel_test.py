import openpyxl
import csv
import math
import os
import time
import requests
import xlrd
import platform
import os.path
import sys

from pathlib import Path

#ERROR_DEFINE
ERROR_STOCKLIST_NOTFOUND = 1
ERROR_STOCKLIST_SHEETNOTFOUND = 2

ERROR_WORKBOOK_SHEETNOTFOUND = 3


RESPONSE_TIMEOUT = 10

FILE_NAME_OUTPUT = '股票.xlsx'
FILE_NAME_STOCK_LIST = 'stocklist.xlsx'

URL= 'http://90.push2his.eastmoney.com/api/qt/stock/kline/get'
HEADER = {
    #'User-Agent': 'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.36',
    #'X-Requested-With': 'XMLHttpRequest'
    'User-Agent: Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/70.0.3538.25 Safari/537.36 Core/1.70.3878.400 QQBrowser/10.8.4518.400',
    'X-Requested-With: XMLHttpRequest'
}

QUERY = {
    #'cb' : 'jQuery112402816109881259974_1638893884237',
    'secid' : '0.000001',
    'ut' : 'fa5fd1943c7b386f172d6893dbfba10b',
    'fields1' : 'f1,f2,f3,f4,f5,f6',
    'fields2' : 'f51,f52,f53,f54,f55,f56,f57,f58,f59,f60,f61',
    'klt' : 103,
    'fqt' : 0,
    'beg' : '20170101',
    'end' : '20500101',
    'smplmt' : 460,
    'lmt' : 1000000,
    '_' : '1638893884280',
}

def test2():
    try:
        r = requests.post(URL, QUERY, HEADER, timeout=RESPONSE_TIMEOUT)
    except Exception as e:
        print(e)
    if r.status_code == requests.codes.ok and r.text != '':
        print('break')
        c = r.json()
        print(c)

    try:
        r.close()
    except Exception as e:
        print(e)

def test():
    wb = openpyxl.load_workbook(FILE_NAME_STOCK_LIST)
    sh = wb['Sheet']
    ce = sh.cell(row = 1,column = 1)   # 读取第一行，第一列的数据
    print(ce.value)
    print(list(sh.rows)[1:])     # 按行读取数据，去掉第一行的表头信息数据
    for cases in list(sh.rows)[1:]:
        case_id =  cases[0].value
        case_excepted = cases[1].value
        case_data = cases[2].value
        print(case_excepted,case_data)
    wb.close()

def copy_stock_code(stockList, excelOutPut):
    stockListSheet = stockList['Sheet']
    if not stockListSheet:
        return ERROR_STOCKLIST_NOTFOUND

    excelOutPutSheet = excelOutPut['Sheet']
    if not excelOutPutSheet:
        return ERROR_WORKBOOK_SHEETNOTFOUND

    max_row = stockListSheet.max_row #最大行数
    max_column = stockListSheet.max_column #最大列数

    for m in range(1, max_row + 1):
        for n in range(97, 97 + max_column): #chr(97)='a'
            n=chr(n)                        #ASCII字符
            i='%s%d'%(n,m)                  #单元格编号
            cell1=stockListSheet[i].value           #获取data单元格数据
            excelOutPutSheet[i].value=cell1           #赋值到test单元格

    excelOutPut.save(FILE_NAME_OUTPUT)

def get_stock_list():
    file = Path(FILE_NAME_STOCK_LIST)
    if not file.is_file():
        return False
    else:
        return openpyxl.load_workbook(FILE_NAME_OUTPUT)

def open_excel():
    file = Path(FILE_NAME_OUTPUT)
    if not file.is_file():
        wb = openpyxl.Workbook()
        wb.save(FILE_NAME_OUTPUT)
    return openpyxl.load_workbook(FILE_NAME_OUTPUT)

def close_excel(wb):
    wb.close()

def open_copy_base_data():
    #load output excel
    file = Path(FILE_NAME_OUTPUT)
    if not file.is_file():
        wb = openpyxl.Workbook()
        wb.save(FILE_NAME_OUTPUT)
    excelOutPut = openpyxl.load_workbook(FILE_NAME_OUTPUT) 

    #local stock list
    file = Path(FILE_NAME_STOCK_LIST)
    if file.is_file():
        stockList = openpyxl.load_workbook(FILE_NAME_STOCK_LIST)  
    else:
        return ERROR_STOCKLIST_NOTFOUND

    #获取sheet页
    sheets1 = excelOutPut.get_sheet_names()
    sheets2 = stockList.get_sheet_names()
    excelOutPutSheet = excelOutPut.get_sheet_by_name(sheets1[0])
    stockListSheet = stockList.get_sheet_by_name(sheets2[0])

    #最大行数
    max_row = stockListSheet.max_row 
    #最大列数
    max_column = stockListSheet.max_column 

    testMap = {}
    for m in range(1, max_row + 1):
        i ='%s%d'%('a', m)
        key = stockListSheet[i].value
        i ='%s%d'%('b', m)
        value = stockListSheet[i].value
        testMap[key] = value

    for x, y in testMap.items():
        print(x, y)

    # for m in range(1, max_row + 1):
    #     for n in range(97, 97 + max_column): #chr(97)='a'
    #         n=chr(n)                        #ASCII字符
    #         i='%s%d'%(n,m)                  #单元格编号
    #         cell1 = stockListSheet[i].value           #获取data单元格数据
    #         excelOutPutSheet[i].value = cell1           #赋值到test单元格
    # excelOutPut.save(FILE_NAME_OUTPUT)
             

if __name__ == '__main__':
    # excelOutPut = open_excel()
    # stockList = get_stock_list()
    # if not stockList:
    #     print('股票列表：stocklist.xlsx文件未找到')
    #     close_excel(excelOutPut)
    #     sys.exit(0)
        
    # copy_stock_code(stockList, excelOutPut)


    # close_excel(excelOutPut)
    # close_excel(stockList)

    ret = open_copy_base_data()
    print(ret)