import openpyxl
import os
import time
import requests
import os.path
import json
import sys
sys.path.append(os.getcwd() + "/script/util")
import util
import string

from pathlib import Path
from util import standardize_dir

def main():
    print('数据爬取完毕，开始处理数据...')
    dataPath = standardize_dir(os.getcwd() + "/financeData/allStockStatement/")
    outputFileName = os.getcwd() + "/股票.xlsx"
    file = Path(outputFileName)
    if not file.is_file():
        wb = openpyxl.Workbook()
        wb.save(outputFileName)

    wb = openpyxl.load_workbook(outputFileName)
    allSheest = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(allSheest[0])
    max_row = ws.max_row         #最大行数
    # max_column = ws.max_column   #最大列数

    iIndex = 0
    for m in range(2, max_row):
        iIndex += 1

        code = ws['%s%d'%('a', m)].value
        fileName = dataPath + str(code)
        file = Path(fileName)
        if not file.is_file():
            continue

        file = open(fileName, 'r')
        data = file.read()
        file.close
        dict = json.loads(data).get("data")

        iSum_TOTALOPERATEREVETZ = 0 # 营业增速
        iCount_TOTALOPERATEREVETZ = 0
        iList = []
        for v in dict:
            iCur = v['TOTALOPERATEREVETZ']
            if not iCur:
                continue

            iCount_TOTALOPERATEREVETZ += 1
            iSum_TOTALOPERATEREVETZ += iCur
            iList.append(iCur)

        # 最新营业额
        # sCell ='%s%d'%('x', m)
        # iSum_TOTALOPERATEREVE = dict[0]['TOTALOPERATEREVE']
        # sReportName = dict[0]['REPORT_TYPE']
        # if iSum_TOTALOPERATEREVE and sReportName:
        #     if sReportName.find("季") > 0:
        #         ws[sCell] = iSum_TOTALOPERATEREVE * 4
        #     else:
        #         ws[sCell] = iSum_TOTALOPERATEREVE
        iSum_TOTALOPERATEREVE = dict[0]['TOTALOPERATEREVE']
        sCell ='%s%d'%('x', m)
        if iSum_TOTALOPERATEREVE:
            ws[sCell] = iSum_TOTALOPERATEREVE   # 最新季度营业额

        # 平均增速
        sCell ='%s%d'%('y', m)
        if iCount_TOTALOPERATEREVETZ == 0:
            ws[sCell] = 0
        else:
            ws[sCell] = iSum_TOTALOPERATEREVETZ / iCount_TOTALOPERATEREVETZ     
        
        sMinCell ='%s%d'%('z', m)
        sMaxCell ='%s%d'%('aa', m)
        if len(iList) > 0:
            ws[sMinCell] = min(iList)   # 最小增速
            ws[sMaxCell] = max(iList)   # 最大增速
        else:
            ws[sMinCell] = 0
            ws[sMaxCell] = 0

        print("正在处理 %s 的数据(%d/%d)"%(code, iIndex, max_row - 1))

    wb.save(outputFileName)
    wb.close
    print('数据全部处理完毕！')

if __name__ == '__main__':
    main()


