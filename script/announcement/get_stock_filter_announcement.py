'''
Author: chanyeow && chanyeow1995@gmail.com
Date: 2022-10-07 15:57:55
LastEditors: chanyeow 767138730@qq.com
LastEditTime: 2022-10-07 23:44:09
Description: 根据特殊字眼，筛选公告选股
Copyright (c) 2022 by chanyeow 767138730@qq.com, All Rights Reserved. 
'''

import openpyxl
import os
import time
import requests
import os.path
import json
import sys

sys.path.append(os.path.dirname(__file__) + "/../util")
import util

from pathlib import Path
from util import get_all_stock_code
from util import get_marketcode
from util import get_market_id

# 爬取的源网站
# https://data.eastmoney.com/notices/hsa/5.html

# 网页版的公告
# https://data.eastmoney.com/notices/detail/601288/AN202209301578824784.html
# 601288 = stock_code, AN202209301578824784 = art_code

# pdf文件地址
# https://pdf.dfcfw.com/pdf/H2_AN202209301578824784_1.pdf?1664577432000.pdf
# AN202209301578824784 = art_code, 1664577432000 = timestamp

# 网页版公告接口
# https://np-cnotice-stock.eastmoney.com/api/content/ann?
# cb=jQuery1123022317735988532816_1665127907444
# &art_code=AN202209301578824784&client_source=web&page_index=1&_=1665127907445

# 获得公告列表的接口
# https://np-anotice-stock.eastmoney.com/api/security/ann?
# cb=jQuery112305847560438381709_1665118651841&sr=-1&page_size=50&page_index=1&ann_type=SHA%2CCYB%2CSZA%2CBJA&client_source=web&f_node=6&s_node=0

#const
FILE_NAME_PDFURL = '/financeData/Announcement_all.xlsx'
MAX_PAGE_COUNT = 1000

TITLE_LIST = ['short_name', 'stock_code', 'art_code', 'notice_date']

KEY_WORDS = [
    "重组",
    "发行股份",
    "支付现金",
    "购买资产",
]

URL= 'https://np-anotice-stock.eastmoney.com/api/security/ann'
HEADER = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/106.0.0.0 Safari/537.36',
    # 'X-Requested-With': 'XMLHttpRequest'
}

# f_node & s_node控制页签
NODE_VALUE = [
    [0, 0], #全部
    [5, 0], [5, 7], [5, 8], [5, 9], #重大事项-全部，重大合同，投资相关，股权激励
    [1, 0], [1, 1], [1, 13], [1, 5], [1, 6], #财务报告-全部，定期报告，利润分配，业绩预告，业绩快报
    [2, 0], [2, 3], [2, 2], [2, 4], #融资公告-全部，增发，新股发行，配股
    [3, 4],     #风险提示
    [6, 0], [6, 10], [6, 11], [6, 12], #资产重组-全部，要约收购，吸收合并，回购
    [4, 12],    #信息变更
    [7, 12],    #持股变动
]

DATA = {
    # 'cb' : 'jQuery11230933578843087203_1665129761622',
    'sr' : -1,
    'page_size' : 50,
    'page_index' : 1,
    'ann_type' : 'SHA,CYB,SZA,BJA',
    'client_source' : 'web',
    'f_node' : 0,   
    's_node' : 0,   
}

def spideringPdfUrl(pageIndex):
    DATA['page_index'] = pageIndex
    ret = False
    try:
        r = requests.get(URL, headers=HEADER, params=DATA)
    except requests.exceptions.RequestException as e:
        print(e)
    r.encoding = 'UTF-8'
    if r.status_code == requests.codes.ok and r.text != '':
        data = r.json()
        ret = True
    else:
        data = {}
        ret = False
    try:
        r.close()
    except Exception as e:
        print(e)
    return data["data"]["list"], ret

def getPdfUrl():
    sFileName = os.getcwd() + FILE_NAME_PDFURL
    file = Path(sFileName)
    if not file.is_file():
        wb = openpyxl.Workbook()
        wb.save(sFileName)
    wb = openpyxl.load_workbook(sFileName)
    allSheest = wb.sheetnames
    ws = wb[allSheest[0]]
    ws.append(TITLE_LIST)

    print('开始爬取公告链接...')
    for pageIndex in range(1, MAX_PAGE_COUNT+1):
        print('尝试爬取第 %d/%d 页...'%(pageIndex, MAX_PAGE_COUNT))
        mData, bRet = spideringPdfUrl(pageIndex)
        if bRet:
            for cell in mData:
                stockData = []
                stockData.append(cell["codes"][0]["short_name"])
                stockData.append(cell["codes"][0]["stock_code"])
                stockData.append(cell["art_code"])
                stockData.append(cell["notice_date"])
                ws.append(stockData)

        time.sleep(0.5)

    wb.save(sFileName)
    wb.close
    print('爬取公告链接完毕...')

def main():
    #getPdfUrl()
    print(NODE_VALUE)

if __name__ == '__main__':
    main()



