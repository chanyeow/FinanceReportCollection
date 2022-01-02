import os
import openpyxl
import os.path

from pathlib import Path

def filter_illegal_string(str):
    illegal_char = {
        ' ': '',
        '*': '',
        '/': '-',
        '\\': '-',
        ':': '-',
        '?': '-',
        '"': '',
        '<': '',
        '>': '',
        '|': '',
        '－': '-',
        '—': '-',
        '（': '(',
        '）': ')',
        'Ａ': 'A',
        'Ｂ': 'B',
        'Ｈ': 'H',
        '，': ',',
        '。': '.',
        '：': '-',
        '！': '_',
        '？': '-',
        '“': '"',
        '”': '"',
        '‘': '',
        '’': ''
    }
    for item in illegal_char.items():
        str = str.replace(item[0], item[1])
    return str

def standardize_dir(dir_str):
    assert (os.path.exists(dir_str)), 'Such directory \"' + str(dir_str) + '\" does not exists!'
    if dir_str[len(dir_str) - 1] != '/':
        return dir_str + '/'
    else:
        return dir_str

def get_all_stock_code():
    mData = {}
    #load file
    filepath = os.getcwd() + '/data/stocklist.xlsx'
    file = Path(filepath)
    if file.is_file():
        stockList = openpyxl.load_workbook(filepath)  
    else:
        return mData

    #get sheet name
    sheets = stockList.get_sheet_names()
    stockListSheet = stockList.get_sheet_by_name(sheets[0])
    #最大行数
    max_row = stockListSheet.max_row 
    #最大列数
    # max_column = stockListSheet.max_column 

    for m in range(1, max_row + 1):
        i ='%s%d'%('a', m)
        key = stockListSheet[i].value
        i ='%s%d'%('b', m)
        value = stockListSheet[i].value
        mData[key] = value
    return mData

def get_code_secid(code):
    if code != '':
        if code[0] == '6':
            marketid = 1
        else:
            marketid = 0
        return '%d.%s'%(marketid, code)
    else:
        return ''

def get_market_id(code):
    if code != '':
        if code[0] == '6':
            return 1
        else:
            return 0
    else:
        return 0

def get_marketcode(code):
    if code != '':
        if code[0] == '6':
            marketName = 'SH'
        else:
            marketName = 'SZ'
        return '%s%s'%(marketName, code)
    else:
        return ''