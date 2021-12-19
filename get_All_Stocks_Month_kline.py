import openpyxl
import math
import os
import time
import requests
import xlrd
import platform
import os.path
import sys
import xml.dom.minidom
import json

from pathlib import Path
from xml.dom.minidom import parse

#ERROR_DEFINE
ERROR_STOCKLIST_NOTFOUND = 1
ERROR_STOCKLIST_SHEETNOTFOUND = 2
ERROR_WORKBOOK_SHEETNOTFOUND = 3

RESPONSE_TIMEOUT = 10

URL= 'http://90.push2his.eastmoney.com/api/qt/stock/kline/get'
HEADER = {
    #'User-Agent': 'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.36',
    #'X-Requested-With': 'XMLHttpRequest'
    'User-Agent: Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/70.0.3538.25 Safari/537.36 Core/1.70.3878.400 QQBrowser/10.8.4518.400',
    'X-Requested-With: XMLHttpRequest'
}

QUERY = {
    #'cb' : 'jQuery112402816109881259974_1638893884237',
    'secid' : '0.000001',
    'ut' : 'fa5fd1943c7b386f172d6893dbfba10b',
    'fields1' : 'f1,f2,f3,f4,f5,f6',
    'fields2' : 'f51,f52,f53,f54,f55,f56,f57,f58,f59,f60,f61',
    'klt' : 103,
    'fqt' : 0,
    'beg' : '20170101',
    'end' : '20500101',
    'smplmt' : 460,
    'lmt' : 1000000,
    '_' : '1638893884280',
}

def get_code_secid(code):
    if code != '':
        if code[0] == '6':
            marketid = 1
        else:
            marketid = 0
        return '%d.%s'%(marketid, code)
    else:
        return ''

FILE_NAME_STOCK_LIST = 'stocklist.xlsx'
def get_all_stock_code():
    mData = {}
    #load file
    file = Path(FILE_NAME_STOCK_LIST)
    if file.is_file():
        stockList = openpyxl.load_workbook(FILE_NAME_STOCK_LIST)  
    else:
        return mData

    #get sheet name
    sheets = stockList.get_sheet_names()
    stockListSheet = stockList.get_sheet_by_name(sheets[0])
    #最大行数
    max_row = stockListSheet.max_row 
    #最大列数
    max_column = stockListSheet.max_column 

    for m in range(1, max_row + 1):
        i ='%s%d'%('a', m)
        key = stockListSheet[i].value
        i ='%s%d'%('b', m)
        value = stockListSheet[i].value
        mData[key] = value
    return mData


PROXIES = {
    # "http": "http://101.132.189.87:9090", 
    "http": "http://101.133.239.96:8080", 
}

def spidering(secid):
    QUERY['secid'] = secid
    try:
        # r = requests.post(URL, QUERY, HEADER, proxies=PROXIES, timeout=RESPONSE_TIMEOUT)
        r = requests.post(URL, QUERY, HEADER, timeout=RESPONSE_TIMEOUT)
    except Exception as e:
        print(e)

    r.encoding = 'utf-8'
    if r.status_code == requests.codes.ok and r.text != '':
        # data = r.json()
        # data =json.loads(r.text).get('data')
        data = r.text
    else:
        data = {}
    try:
        r.close()
    except Exception as e:
        print(e)
    return data

def main():
    mData = get_all_stock_code()
    iTotalCount = len(mData)
    iIndex = 0
    for code, name in mData.items():
        iIndex += 1
        secid = get_code_secid(code)
        allData = spidering(secid)
        
        print("正在爬取 %s - %s 的数据(%d/%d)"%(code, name, iIndex, iTotalCount))
        time.sleep(0.01)

        if iIndex == 1:
            break

if __name__ == '__main__':
    file_name = 'klines.xml'
    last_file_name = 'klines_old.xml'
    main(file_name, last_file_name)
